#########################
#Russell Arlt           #
#CIT 112 David Hosler   #
#PC Specs spreadsheet   #
#########################

from openpyxl import Workbook #importing the workbook module
import pyinputplus as p # pyinputplus for some verification

Monitor_list = [] # a list for monitors!
Graphics_list = [] # a list for graphics cards!
HDD_list = [] # a list for hard drives!
Comp_Dict = {'Case- ' : '', 'Motherboard- ' : '', 'Processor- ' : '', 'Power Supply- ' : '' } # A dictionary for some of the values
Comp_Dict_Lists = {'Graphics Card- ' : Graphics_list, 'Hard Drive- ' : HDD_list, 'Monitor- ' : Monitor_list} # a dictionary for the values stored as lists.
while True: # loop
    user_name = input("Good Morning, what is your name?") # getting user name
    if user_name == '': # can't not have a name
        print("You must have a name.") # yeah
    else: # if they have a name
        print("A lovely name.") # im sure its true
        break # break away

print("We need more computer information! MORE!") # we love that computer info
while True: # loop
    Case_q = input("What kind of a case do you have?") #get case info
    if Case_q == '': # gotta put something
        print("You must have some sort of case") # tell em!
    else: # must be fine
        Comp_Dict['Case- '] = Case_q # update dictionary
        break # break away
while True: # loop
    Mother_q = input("Tell us of your Motherboard.") # get Motherboard info
    if Mother_q == '': # cannot be blank
        print("I know supplies are low but you must have some motherboard.") # true
    else: # all is well
        Comp_Dict['Motherboard- '] = Mother_q # update dictionary
        break # break away
while True: # loop
    Proc_q = input("The Processor that your computer uses is...") # get that info now
    if Proc_q == '': # cannot be empty
        print("You should know what happens if you put nothing in here.") # true
    else: # all good
        Comp_Dict['Processor- '] = Proc_q #update dictionary
        break # break away
while True: # loop
    Pow_q = input("What is your Power Supply?") # get info
    if Pow_q == '' or Pow_q == ' ': # cannot be blank
        print("Enter something please.") # please
    else: # good
        Comp_Dict['Power Supply- '] = Pow_q # update dictionary
        break # break loop
num = 0 # a number to increment
checkycheck = 1 # to know what number graphics card they are on
num_g_c = p.inputInt("How many Graphics cards do you have?") # get the integer input
while True: #while loop to get as many inputs as needed
    if num_g_c > num: #if the input number is greater than num to increment
        Graphics_Card_get = input(f"Please enter a Graphics card {checkycheck}: ") #variable for the input
        if Graphics_Card_get == '' or Graphics_Card_get == ' ': #if they don't put anything in there it kicks them back to the top.
            print("WHOOPSIE DO! You didn't enter anything. Try again.") #a perfect print line.
            continue #kicking them to the top of the while loop
        else: #this will enter anything that they put into the input above.
            Graphics_list.append(Graphics_Card_get) #appends the CPU list with the input asked above.
            num += 1 #increment the number
            checkycheck += 1 #increment the number for the graphics card
            continue #goes to the top of the while loop.
    elif num_g_c <= num: # once it is equal to or less than num
        break # break away
num = 0 # number to increment
checkycheck = 1 # to know what hard drive they are on
num_HDD = p.inputInt("How many Hard Drives do you have?") # get an integer input
while True: #while loop to get as many inputs as needed
    if num_HDD > num: # if the input number is greater than the num to increment
        HDD_get = input(f"Please enter HDD {checkycheck}: ") #variable for the input
        if HDD_get == '' or HDD_get == ' ': #if they don't put anything in there it kicks them back to the top.
            print("WHOOPSIE DO! You didn't enter anything. Try again.") #a perfect print line.
            continue #kicking them to the top of the while loop
        else: #this will enter anything that they put into the input above.
            HDD_list.append(HDD_get) #appends the CPU list with the input asked above.
            num += 1 # increment the number
            checkycheck += 1 # increment the number for the graphics card
            continue #goes to the top of the while loop.
    elif num_HDD <= num: # once it is equal to or less than num
        break # break away
    break # break away
num = 0 # number to increment
checkycheck = 1 # to know what hard drive they are on
num_Monitors = p.inputInt("How many Monitors do you have?") # get an integer input
while True: #while loop to get as many inputs as needed
    if num_Monitors > num: # if the input number is greater than num
        Mon_get = input(f'Please enter Monitor {checkycheck}: ') #variable for the input
        if Mon_get == '' or Mon_get == ' ': #if they don't put anything in there it kicks them back to the top.
            print("WHOOPSIE DO! You didn't enter anything. Try again.") #a perfect print line.
            continue #kicking them to the top of the while loop
        else: #this will enter anything that they put into the input above.
            Monitor_list.append(Mon_get) #appends the CPU list with the input asked above.
            num += 1 # increment the number
            checkycheck += 1 # increment the number for the monitors
            continue #goes to the top of the while loop.
    elif num_Monitors <= num: # once it is equal to or less than num
        break # break out
PC_wb = Workbook() # open a Workbook
ws = PC_wb.active # make sure its the active workbook
ws.title = 'PC Specs' # title the worksheet
ws['A1'] = 'Item Type' # set cell A1 to Item Type
ws['B1'] = 'Item Name' # set Cell B1 to Item Name
ws['C1'] = 'Quantity' # ser cell C1 to Quantity
c_num = 6 # start a variable at 6, you'll see why
var = 0 # an int a 0
for i in range(2, 5): # we know the first 4 quantities are 1, so we set them here
    ws[f'C{i}'] = 1 # setting the cell C2-5 at 1
for i in Graphics_list: # for each thing in the graphics list
    var += 1 # add 1 to var
    c_num += 1 # add one to c_num
ws['C6'] = var # we set C6 to the number we find in var
var = 0 # reset var
for i in HDD_list: # for each thing in the HDD list
    var += 1 # increment var 1
ws[f'C{c_num}'] = var #set the quantity at C c_num to var- the number of things in the list
for i in HDD_list: # look in the HDD list again
    c_num += 1 # increment c_num again
var = 0 # reset var
for i in Monitor_list: # for each thing in the list
    var += 1 # add one to var
ws[f'C{c_num}'] = var #set the cell C c_num to the var in said list.

num = 2 # we have a variable set to 2 here!
for k, v in Comp_Dict.items(): # we look at the keys and values in the dictionary
    ws[f'A{num}'] = k # set cell A num to the key
    ws[f'B{num}'] = v # set cell B num to the value
    num += 1 # increment the num
ws[f'A{num}'] = 'Graphics Card- ' # set cell A num to Graphics Card
for i in Graphics_list: # we look at all items in graphics list
    ws[f'B{num}'] = i # set cell B num to each item
    num += 1 # increment
ws[f'A{num}'] = 'Hard Drive- ' #set cell A num to Hard Drive
for i in HDD_list: # for each one in the list
    ws[f'B{num}'] = i # set cell B num to it
    num += 1 # increment
ws[f'A{num}'] = 'Monitor- ' # set cell A num to Monitor
for i in Monitor_list: # for each in Monitor list
    ws[f'B{num}'] = i # set cell B num to it
    num += 1 # increment

PC_wb.save(f'{user_name}_pcspecs.xlsx') # Save or overwrite spreadsheet
